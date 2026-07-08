"""
step8_fixed_final.py
====================
Ολοκληρωμένο module για Fill + Optimize (χωρίς Streamlit UI).

Παρέχει όλη τη λειτουργικότητα του app.py ως library/CLI:
- Phase 1: Fill template με δεδομένα μαθητών
- Phase 2: Optimization με asymmetric swaps
- Locked students support (ΖΩΗΡΟΣ, ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ, ΙΔΙΑΙΤΕΡΟΤΗΤΑ)
- ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ + SINGLE sheets
- Detailed statistics + swaps log

Απαιτήσεις: openpyxl>=3.1.0
"""
from __future__ import annotations

import sys
import re
import unicodedata
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet


# ========== DATACLASSES (FIXED για Python 3.12) ==========

@dataclass
class StudentData:
    """Δεδομένα μαθητή από source (Phase 1)."""
    name: str = ""
    gender: str = ""
    teacher_child: str = "Ο"
    calm: str = "Ο"
    special_needs: str = "Ο"
    greek_knowledge: str = "Ν"
    friends: List[str] = field(default_factory=list)
    conflict_names: List[str] = field(default_factory=list)  # δηλωμένες συγκρούσεις
    choice: int = 1


@dataclass
class Student:
    """Student για optimizer (Phase 2)."""
    name: str = ""
    choice: int = 1
    gender: str = ""
    greek_knowledge: str = "Ν"
    friends: List[str] = field(default_factory=list)
    conflict_names: List[str] = field(default_factory=list)  # δηλωμένες συγκρούσεις
    locked: bool = False


class LegacyExcelFormatError(ValueError):
    """Το input Excel δεν περιέχει τις νέες raw στήλες (ΦΥΛΟ_Α/Β, ΓΝΩΣΗ_Α/Β,
    ΣΥΓΚΡΟΥΣΗ_Α/Β, ΣΥΓΚΡΟΥΣΗ) που παράγει το τρέχον 'fill' mode.

    Σηκώνεται fail-fast από το load_filled_data ώστε να ΜΗΝ τρέξει
    optimization πάνω σε σιωπηλά ελλιπή δεδομένα (default gender='Α',
    greek='Ν', καμία σύγκρουση). Αν το legacy fallback είναι επιθυμητό,
    πέρασε ρητά allow_legacy_fallback=True.
    """




# ========== SHARED NAME HELPERS ==========

def _strip_diacritics(text: str) -> str:
    nfkd = unicodedata.normalize("NFD", str(text))
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch))


def _canon_name(x: Any) -> str:
    """Ενιαία κανονικοποίηση ονομάτων για φίλους/συγκρούσεις."""
    if x is None:
        return ""
    s = str(x).strip().strip("[]'\" ")
    s = re.sub(r"\s+", " ", s)
    return _strip_diacritics(s).upper()


def _split_name_list(value: Any) -> List[str]:
    """Ασφαλές split λίστας ονομάτων από Excel cell."""
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        raw = list(value)
    else:
        s = str(value).strip()
        if not s or s.upper() in {"NAN", "NONE", "NULL", "-"}:
            return []
        raw = re.split(r"[,;|/·\n]+", s)
    return [str(x).strip() for x in raw if str(x).strip()]


def _name_in_list(target: str, names: List[str]) -> bool:
    target_c = _canon_name(target)
    return bool(target_c) and target_c in {_canon_name(n) for n in names}


def _are_mutual_friends(name_a: str, friends_a: List[str], name_b: str, friends_b: List[str]) -> bool:
    """True μόνο για πλήρως αμοιβαία δυάδα."""
    return _name_in_list(name_b, friends_a) and _name_in_list(name_a, friends_b)

# ========== MAIN PROCESSOR CLASS ==========

class UnifiedProcessor:
    """Ενοποιημένος processor: Fill + Optimize."""
    
    def __init__(self):
        self.students_data: Dict[str, StudentData] = {}
        self.teams_students: Dict[str, List[str]] = {}
        self.students: Dict[str, Student] = {}
        self.teams: Dict[str, List[str]] = {}
        self.target_ep3 = 3
        self.target_gender = 4
        self.target_greek = 4
        self.warnings: List[str] = []
    
    # ==================== PHASE 1: FILL ====================
    
    def read_source_data(self, source_path: str) -> None:
        """Διάβασμα δεδομένων από Παράδειγμα1.xlsx."""
        wb = load_workbook(source_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = self._parse_headers_fill(sheet)
            
            if 'ΟΝΟΜΑ' not in headers:
                continue
            
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                
                if not name:
                    continue
                
                # Friends
                friends_str = self._get_cell_value(sheet, row_idx, headers.get('ΦΙΛΟΙ'))
                friends = _split_name_list(friends_str)

                # Conflicts
                conflict_str = self._get_cell_value(sheet, row_idx, headers.get('ΣΥΓΚΡΟΥΣΗ'))
                conflict_names = _split_name_list(conflict_str)
                
                # Choice (ΕΠΙΔΟΣΗ)
                choice_val = 1
                if 'ΕΠΙΔΟΣΗ' in headers:
                    epidosi_raw = sheet.cell(row_idx, headers['ΕΠΙΔΟΣΗ']).value
                    if epidosi_raw is not None:
                        try:
                            choice_val = int(epidosi_raw)
                        except:
                            choice_val = 1
                
                # Greek knowledge - robust parsing
                greek_raw = None
                found_greek = False
                for variant in ['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΚΑΛΗ ΓΝΩΣΗ ΕΛΛΗΝΙΚΩΝ', 
                               'ΚΑΛΗ_ΓΝΩΣΗ', 'ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ']:
                    if variant in headers:
                        greek_raw = self._get_cell_value(sheet, row_idx, headers[variant], None)
                        if greek_raw is not None and greek_raw != '':
                            found_greek = True
                            break
                
                # Process Greek knowledge
                if not found_greek or greek_raw is None or greek_raw == '':
                    self.warnings.append(f"⚠️ Μαθητής {name}: Δεν βρέθηκε ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ - παραλείπεται")
                    continue
                else:
                    greek_str = str(greek_raw).strip().upper()
                    if greek_str.startswith('Ν') or greek_str.startswith('N'):
                        greek_val = 'Ν'
                    elif greek_str.startswith('Ο') or greek_str.startswith('O'):
                        greek_val = 'Ο'
                    else:
                        self.warnings.append(f"⚠️ Unknown ΚΑΛΗ_ΓΝΩΣΗ '{greek_raw}' for {name}, defaulting to Ν")
                        greek_val = 'Ν'
                
                self.students_data[name] = StudentData(
                    name=name,
                    gender=self._get_cell_value(sheet, row_idx, headers.get('ΦΥΛΟ'), 'Κ'),
                    teacher_child=self._get_cell_value(sheet, row_idx, headers.get('ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'), 'Ο'),
                    calm=self._get_cell_value(sheet, row_idx, headers.get('ΖΩΗΡΟΣ'), 'Ο'),
                    special_needs=self._get_cell_value(sheet, row_idx, headers.get('ΙΔΙΑΙΤΕΡΟΤΗΤΑ'), 'Ο'),
                    greek_knowledge=greek_val,
                    friends=friends,
                    conflict_names=conflict_names,
                    choice=choice_val
                )
        
        wb.close()
        print(f"✅ Διαβάστηκαν {len(self.students_data)} μαθητές από source file")
    
    def fill_target_excel(self, template_path: str, output_path: str) -> str:
        """Συμπλήρωση STEP7_TEMPLATE."""
        wb = load_workbook(template_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            filled_count = self._fill_sheet(sheet, sheet_name)
            if filled_count > 0:
                print(f"📝 Sheet '{sheet_name}': {filled_count} μαθητές")
        
        self._create_categorization_sheet(wb)
        
        wb.save(output_path)
        wb.close()
        
        print(f"✅ Filled Excel αποθηκεύτηκε: {output_path}")
        return output_path
    
    def _fill_sheet(self, sheet: Worksheet, team_name: str) -> int:
        """Συμπλήρωση ενός sheet."""
        headers_map = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                header = str(cell.value).strip().upper()
                header_key = header.replace('_', '').replace(' ', '')
                headers_map[header_key] = col_idx
        
        if 'ΟΝΟΜΑ' not in headers_map:
            return 0
        
        required_headers = ['ΦΥΛΟ', 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
        next_col = max(headers_map.values()) + 1
        
        for req_header in required_headers:
            if req_header not in headers_map:
                cell = sheet.cell(1, next_col)
                original_header = req_header.replace('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ')
                cell.value = original_header
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                headers_map[req_header] = next_col
                next_col += 1
        
        filled_count = 0
        
        if team_name not in self.teams_students:
            self.teams_students[team_name] = []
        
        for row_idx in range(2, sheet.max_row + 1):
            name_cell = sheet.cell(row_idx, headers_map['ΟΝΟΜΑ'])
            name = name_cell.value
            
            if not name or str(name).strip() == '':
                continue
            
            name = str(name).strip()
            
            if name not in self.students_data:
                continue
            
            student_data = self.students_data[name]
            self.teams_students[team_name].append(name)
            
            # Fill ΦΥΛΟ
            if 'ΦΥΛΟ' in headers_map:
                col = headers_map['ΦΥΛΟ']
                sheet.cell(row_idx, col).value = student_data.gender
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            # Fill ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ
            if 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ' in headers_map:
                col = headers_map['ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ']
                sheet.cell(row_idx, col).value = student_data.greek_knowledge
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            # Fill ΦΙΛΟΙ
            if 'ΦΙΛΟΙ' in headers_map:
                col = headers_map['ΦΙΛΟΙ']
                sheet.cell(row_idx, col).value = ', '.join(student_data.friends) if student_data.friends else ''
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='left', vertical='center')
            
            # Fill ΕΠΙΔΟΣΗ
            if 'ΕΠΙΔΟΣΗ' in headers_map:
                col = headers_map['ΕΠΙΔΟΣΗ']
                sheet.cell(row_idx, col).value = student_data.choice
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            filled_count += 1
        
        return filled_count
    
    def _create_categorization_sheet(self, workbook: Workbook) -> None:
        """Δημιουργία sheet ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ."""
        if 'ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ' in workbook.sheetnames:
            del workbook['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ']
        
        cat_sheet = workbook.create_sheet('ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ')

        # FIX: προστέθηκαν raw στήλες (ΦΥΛΟ/ΓΝΩΣΗ/ΣΥΓΚΡΟΥΣΗ ανά μαθητή) ώστε το
        # sheet να είναι αυτοτελές (self-contained) και το standalone 'optimize'
        # mode να μη χρειάζεται πλέον το αρχικό source file (self.students_data).
        headers = ['ΜΑΘΗΤΗΣ Α', 'ΜΑΘΗΤΗΣ Β', 'ΚΑΤΗΓΟΡΙΑ ΔΥΑΔΑΣ', 'ΕΠΙΔΟΣΗ', 'LOCKED', 'ΤΜΗΜΑ',
                   'ΦΥΛΟ_Α', 'ΓΝΩΣΗ_Α', 'ΣΥΓΚΡΟΥΣΗ_Α', 'ΦΥΛΟ_Β', 'ΓΝΩΣΗ_Β', 'ΣΥΓΚΡΟΥΣΗ_Β']
        for col_idx, header in enumerate(headers, start=1):
            cell = cat_sheet.cell(1, col_idx)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        all_students = []
        for team_name in sorted(self.teams_students.keys()):
            for student_name in self.teams_students[team_name]:
                if student_name in self.students_data:
                    student = self.students_data[student_name]
                    all_students.append({
                        'name': student_name,
                        'data': student,
                        'team': team_name
                    })
        
        row_idx = 2
        processed = set()
        
        for i, student_a in enumerate(all_students):
            if student_a['name'] in processed:
                continue
            
            for j, student_b in enumerate(all_students[i+1:], start=i+1):
                if student_b['name'] in processed:
                    continue
                
                if _are_mutual_friends(
                    student_a['name'], student_a['data'].friends,
                    student_b['name'], student_b['data'].friends
                ):
                    
                    category = self._determine_category(
                        student_a['data'].gender,
                        student_a['data'].greek_knowledge,
                        student_b['data'].gender,
                        student_b['data'].greek_knowledge
                    )
                    
                    epidosi_text = f"{student_a['data'].choice}, {student_b['data'].choice}"
                    
                    cat_sheet.cell(row_idx, 1).value = student_a['name']
                    cat_sheet.cell(row_idx, 2).value = student_b['name']
                    cat_sheet.cell(row_idx, 3).value = category
                    cat_sheet.cell(row_idx, 4).value = epidosi_text
                    
                    is_locked = (self._is_student_locked(student_a['data']) or 
                                 self._is_student_locked(student_b['data']))
                    cat_sheet.cell(row_idx, 5).value = 'LOCKED' if is_locked else 'ΟΧΙ'
                    
                    if is_locked:
                        team_text = 'LOCKED'
                    else:
                        team_text = f"{student_a['team']},{student_b['team']}"
                    cat_sheet.cell(row_idx, 6).value = team_text

                    # FIX: γράφουμε raw δεδομένα ώστε να ξαναδιαβάζονται 1-προς-1
                    # στο load_filled_data χωρίς εξάρτηση από self.students_data.
                    cat_sheet.cell(row_idx, 7).value = student_a['data'].gender
                    cat_sheet.cell(row_idx, 8).value = student_a['data'].greek_knowledge
                    cat_sheet.cell(row_idx, 9).value = ', '.join(student_a['data'].conflict_names)
                    cat_sheet.cell(row_idx, 10).value = student_b['data'].gender
                    cat_sheet.cell(row_idx, 11).value = student_b['data'].greek_knowledge
                    cat_sheet.cell(row_idx, 12).value = ', '.join(student_b['data'].conflict_names)

                    for col in range(1, 13):
                        cat_sheet.cell(row_idx, col).alignment = Alignment(
                            horizontal='left' if col <= 2 else 'center',
                            vertical='center'
                        )
                    
                    processed.add(student_a['name'])
                    processed.add(student_b['name'])
                    row_idx += 1
                    break
        
        cat_sheet.column_dimensions['A'].width = 30
        cat_sheet.column_dimensions['B'].width = 30
        cat_sheet.column_dimensions['C'].width = 35
        cat_sheet.column_dimensions['D'].width = 12
        cat_sheet.column_dimensions['E'].width = 12
        cat_sheet.column_dimensions['F'].width = 20
        for col in ['G', 'H', 'I', 'J', 'K', 'L']:
            cat_sheet.column_dimensions[col].width = 14

        self._create_single_sheet(workbook, all_students, processed)
    
    def _is_student_locked(self, student: StudentData) -> bool:
        """Έλεγχος αν μαθητής είναι locked."""
        return (student.calm == 'Ν' or 
                student.teacher_child == 'Ν' or 
                student.special_needs == 'Ν')
    
    def _determine_category(self, gender_a: str, greek_a: str, gender_b: str, greek_b: str) -> str:
        """Καθορισμός κατηγορίας δυάδας."""
        if gender_a != gender_b:
            return "Ομάδες Μικτού Φύλου"
        
        gender_label = "Κορίτσια" if gender_a == "Κ" else "Αγόρια"
        
        if greek_a == greek_b:
            if greek_a == "Ν":
                return f"Καλή Γνώση ({gender_label})"
            else:
                return f"όχι Καλή Γνώση ({gender_label})"
        else:
            return f"Μικτής Γνώσης ({gender_label})"
    
    def _determine_single_category(self, gender: str, greek_knowledge: str) -> str:
        """Καθορισμός κατηγορίας για μεμονωμένο μαθητή."""
        gender_label = "Κορίτσια" if gender == "Κ" else "Αγόρια"
        
        if greek_knowledge == "Ν":
            return f"{gender_label} - Ν (Καλή γνώση)"
        else:
            return f"{gender_label} - Ο (όχι καλή γνώση)"
    
    def _create_single_sheet(self, workbook: Workbook, all_students: List[Dict], processed_names: set) -> None:
        """Δημιουργία sheet SINGLE."""
        if 'SINGLE' in workbook.sheetnames:
            del workbook['SINGLE']
        
        single_sheet = workbook.create_sheet('SINGLE')
        
        # FIX: προστέθηκε στήλη ΣΥΓΚΡΟΥΣΗ ώστε το SINGLE sheet να είναι επίσης
        # αυτοτελές (self-contained) όπως το ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ.
        headers = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΚΑΤΗΓΟΡΙΑ SINGLE',
                   'ΤΜΗΜΑ', 'LOCKED', 'ΣΥΓΚΡΟΥΣΗ']
        for col_idx, header in enumerate(headers, start=1):
            cell = single_sheet.cell(1, col_idx)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        single_students = []
        for student in all_students:
            if student['name'] not in processed_names:
                single_students.append(student)
        
        single_students.sort(key=lambda x: x['name'])
        
        for row_idx, student in enumerate(single_students, start=2):
            student_data = student['data']
            team_name = student['team']
            
            single_sheet.cell(row_idx, 1).value = student['name']
            single_sheet.cell(row_idx, 2).value = student_data.gender
            single_sheet.cell(row_idx, 3).value = student_data.greek_knowledge
            single_sheet.cell(row_idx, 4).value = student_data.choice
            
            category = self._determine_single_category(student_data.gender, student_data.greek_knowledge)
            single_sheet.cell(row_idx, 5).value = category
            
            is_locked = self._is_student_locked(student_data)
            single_sheet.cell(row_idx, 7).value = 'LOCKED' if is_locked else 'ΟΧΙ'
            
            if is_locked:
                single_sheet.cell(row_idx, 6).value = 'LOCKED'
            else:
                single_sheet.cell(row_idx, 6).value = team_name

            single_sheet.cell(row_idx, 8).value = ', '.join(student_data.conflict_names)

            for col in range(1, 9):
                single_sheet.cell(row_idx, col).alignment = Alignment(
                    horizontal='left' if col == 1 else 'center',
                    vertical='center'
                )
        
        single_sheet.column_dimensions['A'].width = 30
        single_sheet.column_dimensions['B'].width = 12
        single_sheet.column_dimensions['C'].width = 25
        single_sheet.column_dimensions['D'].width = 12
        single_sheet.column_dimensions['E'].width = 35
        single_sheet.column_dimensions['F'].width = 20
        single_sheet.column_dimensions['G'].width = 12
        single_sheet.column_dimensions['H'].width = 25
    
    # ==================== PHASE 2: OPTIMIZE ====================
    
    def load_filled_data(self, filled_path: str, allow_legacy_fallback: bool = False) -> None:
        """Φόρτωση δεδομένων από filled Excel για optimization.

        Args:
            filled_path: Path στο filled Excel (output του 'fill' mode).
            allow_legacy_fallback: Αν False (default), σηκώνει
                ``LegacyExcelFormatError`` fail-fast όταν λείπουν οι νέες raw
                στήλες, αντί να κάνει σιωπηλό fallback σε defaults. Πέρασέ το
                True μόνο αν αποδέχεσαι ρητά ελλιπή δεδομένα φύλου/γνώσης/
                συγκρούσεων για παλιά αρχεία.

        Raises:
            LegacyExcelFormatError: αν λείπουν οι νέες στήλες και
                allow_legacy_fallback είναι False.
        """
        wb = load_workbook(filled_path, data_only=True)

        self._validate_new_schema(wb, allow_legacy_fallback)

        if 'ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ' in wb.sheetnames:
            self._load_from_kategoriopoihsh(wb['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ'])
        
        if 'SINGLE' in wb.sheetnames:
            self._load_from_single(wb['SINGLE'])
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ', 'SINGLE']:
                continue
            
            sheet = wb[sheet_name]
            headers = self._parse_headers(sheet)
            
            if 'ΟΝΟΜΑ' not in headers:
                continue
            
            self.teams[sheet_name] = []
            
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                if name and name in self.students:
                    self.teams[sheet_name].append(name)
        
        wb.close()
        print(f"✅ Φορτώθηκαν {len(self.students)} students, {len(self.teams)} teams")
    
    def _validate_new_schema(self, wb: Workbook, allow_legacy_fallback: bool) -> None:
        """Ελέγχει ρητά (fail-fast) το σχήμα στηλών πριν τρέξει οποιαδήποτε
        λογική optimization. Αντικαθιστά το προηγούμενο 'σιωπηλό warning +
        fallback': τώρα, αν λείπουν οι νέες στήλες, σταματάει αμέσως με σαφές
        μήνυμα, εκτός αν το caller περάσει ρητά allow_legacy_fallback=True.
        """
        problems: List[str] = []

        if 'ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ' in wb.sheetnames:
            headers = self._parse_headers(wb['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ'])
            required_new = ['ΦΥΛΟΑ', 'ΓΝΩΣΗΑ', 'ΦΥΛΟΒ', 'ΓΝΩΣΗΒ']
            missing = [h for h in required_new if h not in headers]
            if missing:
                problems.append(f"ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ: λείπουν στήλες {missing}")

        if 'SINGLE' in wb.sheetnames:
            headers = self._parse_headers(wb['SINGLE'])
            if 'ΣΥΓΚΡΟΥΣΗ' not in headers:
                problems.append("SINGLE: λείπει στήλη ΣΥΓΚΡΟΥΣΗ")

        if not problems:
            return

        message = (
            "Το input Excel φαίνεται να έχει παραχθεί από ΠΑΛΙΟ 'fill' mode "
            "(λείπουν νέες στήλες φύλου/γνώσης/σύγκρουσης): "
            + "; ".join(problems)
            + ". Ξανάτρεξε το mode 'fill' στο σωστό source για πλήρη "
              "δεδομένα, ή πέρασε ρητά allow_legacy_fallback=True "
              "(CLI: --allow-legacy-fallback) αν αποδέχεσαι ελλιπή δεδομένα "
              "(defaults: gender='Α', greek='Ν', καμία σύγκρουση)."
        )

        if allow_legacy_fallback:
            self.warnings.append(f"⚠️ {message}")
        else:
            raise LegacyExcelFormatError(message)

    def _load_from_kategoriopoihsh(self, sheet: Worksheet) -> None:
        """Διάβασμα δυάδων από ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ sheet.

        FIX: πριν, το φύλο/γνώση/συγκρούσεις διαβάζονταν από
        ``self.students_data`` (in-memory cache που γεμίζει μόνο η
        ``read_source_data``). Στο standalone CLI mode ``optimize`` η cache
        αυτή είναι πάντα άδεια, οπότε ΟΛΟΙ οι μαθητές έπαιρναν σιωπηλά τις
        default τιμές ('Α', 'Ν', χωρίς locked/conflicts) — λάθος αποτελέσματα
        χωρίς κανένα σφάλμα/warning. Τώρα διαβάζουμε πρώτα τις raw στήλες
        ΦΥΛΟ_Α/ΓΝΩΣΗ_Α/ΣΥΓΚΡΟΥΣΗ_Α (Β αντίστοιχα) που γράφει η
        ``_create_categorization_sheet``· αν λείπουν (παλιό αρχείο χωρίς τις
        στήλες αυτές), γίνεται fallback σε ``self.students_data`` με warning.
        """
        headers = self._parse_headers(sheet)
        
        required = ['ΜΑΘΗΤΗΣΑ', 'ΜΑΘΗΤΗΣΒ', 'ΚΑΤΗΓΟΡΙΑΔΥΑΔΑΣ', 'ΕΠΙΔΟΣΗ']
        missing = [h for h in required if h not in headers]
        if missing:
            return

        has_raw_cols = all(
            h in headers for h in ['ΦΥΛΟΑ', 'ΓΝΩΣΗΑ', 'ΦΥΛΟΒ', 'ΓΝΩΣΗΒ']
        )
        if not has_raw_cols:
            self.warnings.append(
                "⚠️ ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ: λείπουν οι στήλες ΦΥΛΟ_Α/ΓΝΩΣΗ_Α/ΦΥΛΟ_Β/"
                "ΓΝΩΣΗ_Β (παλιό αρχείο) — γίνεται fallback σε "
                "self.students_data (χρειάζεται προηγούμενο read_source_data)."
            )

        for row_idx in range(2, sheet.max_row + 1):
            name_a = self._get_cell_value(sheet, row_idx, headers.get('ΜΑΘΗΤΗΣΑ'))
            name_b = self._get_cell_value(sheet, row_idx, headers.get('ΜΑΘΗΤΗΣΒ'))
            category = self._get_cell_value(sheet, row_idx, headers.get('ΚΑΤΗΓΟΡΙΑΔΥΑΔΑΣ'))
            epidosh_raw = self._get_cell_value(sheet, row_idx, headers.get('ΕΠΙΔΟΣΗ'))
            
            if not name_a or not name_b or not category:
                continue
            
            # Parse επίδοση
            epidosh_a, epidosh_b = 1, 1
            if ',' in epidosh_raw:
                parts = epidosh_raw.split(',')
                try:
                    epidosh_a = int(parts[0].strip())
                    epidosh_b = int(parts[1].strip())
                except (ValueError, IndexError):
                    pass

            if has_raw_cols:
                gender_a = self._get_cell_value(sheet, row_idx, headers.get('ΦΥΛΟΑ'), 'Α')
                gender_b = self._get_cell_value(sheet, row_idx, headers.get('ΦΥΛΟΒ'), 'Α')
                greek_a = self._get_cell_value(sheet, row_idx, headers.get('ΓΝΩΣΗΑ'), 'Ν')
                greek_b = self._get_cell_value(sheet, row_idx, headers.get('ΓΝΩΣΗΒ'), 'Ν')
                conflicts_a_raw = self._get_cell_value(sheet, row_idx, headers.get('ΣΥΓΚΡΟΥΣΗΑ'))
                conflicts_b_raw = self._get_cell_value(sheet, row_idx, headers.get('ΣΥΓΚΡΟΥΣΗΒ'))
                conflicts_a = _split_name_list(conflicts_a_raw)
                conflicts_b = _split_name_list(conflicts_b_raw)
                locked_val = self._get_cell_value(sheet, row_idx, headers.get('LOCKED'))
                is_locked = (locked_val == 'LOCKED')
            else:
                # Fallback (backward-compatible) στο παλιό, εξαρτημένο από source
                sa = self.students_data.get(name_a)
                sb = self.students_data.get(name_b)
                gender_a = sa.gender if sa else 'Α'
                gender_b = sb.gender if sb else 'Α'
                greek_a = sa.greek_knowledge if sa else 'Ν'
                greek_b = sb.greek_knowledge if sb else 'Ν'
                conflicts_a = sa.conflict_names if sa else []
                conflicts_b = sb.conflict_names if sb else []
                is_locked = (self._is_student_locked(sa) if sa else False) or \
                            (self._is_student_locked(sb) if sb else False)

            if name_a not in self.students:
                self.students[name_a] = Student(
                    name=name_a,
                    choice=epidosh_a,
                    gender=gender_a,
                    greek_knowledge=greek_a,
                    friends=[name_b],
                    conflict_names=conflicts_a,
                    locked=is_locked
                )
            
            if name_b not in self.students:
                self.students[name_b] = Student(
                    name=name_b,
                    choice=epidosh_b,
                    gender=gender_b,
                    greek_knowledge=greek_b,
                    friends=[name_a],
                    conflict_names=conflicts_b,
                    locked=is_locked
                )
    
    def _load_from_single(self, sheet: Worksheet) -> None:
        """Διάβασμα μονών μαθητών από SINGLE sheet."""
        headers = self._parse_headers(sheet)
        
        required = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ']
        missing = [h for h in required if h not in headers]
        if missing:
            return
        
        for row_idx in range(2, sheet.max_row + 1):
            name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
            if not name:
                continue
            
            if name in self.students:
                continue
            
            gender_col = headers.get('ΦΥΛΟ')
            greek_col = (headers.get('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ') or 
                        headers.get('ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ') or
                        headers.get('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ'))
            epidosh_col = headers.get('ΕΠΙΔΟΣΗ')
            locked_col = headers.get('LOCKED')
            
            gender = self._get_cell_value(sheet, row_idx, gender_col, 'Α')
            
            # Greek knowledge
            raw_greek = sheet.cell(row_idx, greek_col).value if greek_col else 'Ν'
            if raw_greek:
                greek_str = str(raw_greek).strip().upper()
                if greek_str.startswith('Ν') or greek_str.startswith('N'):
                    greek = 'Ν'
                elif greek_str.startswith('Ο') or greek_str.startswith('O'):
                    greek = 'Ο'
                else:
                    greek = 'Ν'
            else:
                greek = 'Ν'
            
            raw_epidosh = sheet.cell(row_idx, epidosh_col).value if epidosh_col else 1
            try:
                epidosh = int(raw_epidosh) if raw_epidosh else 1
            except:
                epidosh = 1
            
            locked_val = self._get_cell_value(sheet, row_idx, locked_col)
            is_locked = (locked_val == 'LOCKED')

            # FIX: πριν έπαιρνε conflicts μόνο από self.students_data (άδειο σε
            # standalone optimize mode). Τώρα διαβάζει πρώτα τη στήλη
            # ΣΥΓΚΡΟΥΣΗ του ίδιου SINGLE sheet, με fallback στην cache για
            # συμβατότητα με παλιά αρχεία.
            conflict_col = headers.get('ΣΥΓΚΡΟΥΣΗ')
            if conflict_col is not None:
                conflicts_raw = self._get_cell_value(sheet, row_idx, conflict_col)
                conflict_names = _split_name_list(conflicts_raw)
            else:
                conflict_names = (
                    self.students_data[name].conflict_names
                    if name in self.students_data else []
                )

            self.students[name] = Student(
                name=name,
                choice=epidosh,
                gender=gender,
                greek_knowledge=greek,
                friends=[],
                conflict_names=conflict_names,
                locked=is_locked
            )
    
    def calculate_spreads(self) -> Dict[str, int]:
        """Υπολογισμός spreads για επίδοση 5→1→4→2, ελληνικά και φύλο."""
        stats = self._get_team_stats()
        if not stats:
            return {
                'ep5': 0, 'ep1': 0, 'ep4': 0, 'ep2': 0, 'ep3': 0,
                'boys': 0, 'girls': 0, 'greek_yes': 0
            }

        def _spread(key: str) -> int:
            vals = [s[key] for s in stats.values()]
            return max(vals) - min(vals) if vals else 0

        return {
            # Ιεραρχία επίδοσης για swaps: 5 → 1 → 4 → 2
            'ep5': _spread('ep5'),
            'ep1': _spread('ep1'),
            'ep4': _spread('ep4'),
            'ep2': _spread('ep2'),
            # Κρατάμε και ep3 για αναφορά/συμβατότητα, αλλά δεν είναι προτεραιότητα swap.
            'ep3': _spread('ep3'),
            'boys': _spread('boys'),
            'girls': _spread('girls'),
            'greek_yes': _spread('greek_yes')
        }

    def _get_team_stats(self) -> Dict:
        """Μέτρηση stats ανά τμήμα."""
        stats = {}
        for team_name, student_names in self.teams.items():
            boys = girls = greek_yes = greek_no = 0
            ep1 = ep2 = ep3 = ep4 = ep5 = 0

            for name in student_names:
                if name not in self.students:
                    continue
                s = self.students[name]

                if s.gender == 'Α':
                    boys += 1
                elif s.gender == 'Κ':
                    girls += 1

                if s.greek_knowledge in ['Ν', 'N']:
                    greek_yes += 1
                elif s.greek_knowledge in ['Ο', 'O']:
                    greek_no += 1

                if s.choice == 1:
                    ep1 += 1
                elif s.choice == 2:
                    ep2 += 1
                elif s.choice == 3:
                    ep3 += 1
                elif s.choice == 4:
                    ep4 += 1
                elif s.choice == 5:
                    ep5 += 1

            stats[team_name] = {
                'boys': boys, 'girls': girls,
                'greek_yes': greek_yes, 'greek_no': greek_no,
                'ep1': ep1, 'ep2': ep2, 'ep3': ep3, 'ep4': ep4, 'ep5': ep5
            }

        return stats

    # ---- Σταθερά ορίων (MAX_SPREAD) ----
    MAX_SPREAD_EPIDOSIS = 2
    MAX_SPREAD_GREEK    = 2
    MAX_SPREAD_GENDER   = 2

    PERFORMANCE_PRIORITY = ('ep5', 'ep1', 'ep4', 'ep2')
    SWAP_PRIORITY_ORDER = ('ep5', 'ep1', 'ep4', 'ep2', 'greek_yes', 'gender')

    def optimize(self, max_iterations: int = 200) -> Tuple[List[Dict], Dict]:
        """
        Ιεραρχημένη πολυκριτηριακή βελτιστοποίηση με swaps.

        Hard rules:
        - Δεν δημιουργεί δηλωμένη ΣΥΓΚΡΟΥΣΗ.
        - Δεν μετακινεί locked μαθητές.
        - Δεν σπάει πλήρως αμοιβαίες δυάδες.

        Ιεραρχία βελτίωσης:
        1. Επίδοση 5
        2. Επίδοση 1
        3. Επίδοση 4
        4. Επίδοση 2
        5. Καλή γνώση ελληνικών
        6. Φύλο
        """
        MAX_EP  = self.MAX_SPREAD_EPIDOSIS
        MAX_GR  = self.MAX_SPREAD_GREEK
        MAX_GEN = self.MAX_SPREAD_GENDER

        applied_swaps = []

        print(f"🔄 Ξεκινά ιεραρχημένη βελτιστοποίηση (max {max_iterations} iterations)...")
        print(f"   Ιεραρχία επίδοσης: 5 → 1 → 4 → 2 | Όρια: επίδοση≤{MAX_EP}, γλώσσα≤{MAX_GR}, φύλο≤{MAX_GEN}")

        for iteration in range(max_iterations):
            all_candidate_swaps = self._generate_all_valid_swaps(MAX_EP, MAX_GR, MAX_GEN)

            if not all_candidate_swaps:
                print(f"✅ Δεν υπάρχουν άλλες νόμιμες βελτιωτικές ανταλλαγές στο iteration {iteration}. Βελτιστοποίηση ολοκληρώθηκε.")
                break

            best_swap = self._select_best_swap_hierarchical(all_candidate_swaps)

            if not best_swap:
                print(f"✅ Δεν βρέθηκε καλύτερη ανταλλαγή στο iteration {iteration}.")
                break

            self._apply_swap(best_swap)
            applied_swaps.append(best_swap)

            spreads = self.calculate_spreads()
            if (iteration + 1) % 10 == 0:
                print(f"  Iteration {iteration + 1}: {len(applied_swaps)} swaps | "
                      f"EP5={spreads['ep5']} EP1={spreads['ep1']} "
                      f"EP4={spreads['ep4']} EP2={spreads['ep2']} "
                      f"GR={spreads['greek_yes']} B={spreads['boys']} G={spreads['girls']}")
        else:
            print(f"⚠️ Έφτασε το όριο των {max_iterations} iterations.")

        final_spreads = self.calculate_spreads()
        print(f"✅ Optimization ολοκληρώθηκε: {len(applied_swaps)} swaps")
        print(f"   Final spreads: EP5={final_spreads['ep5']}, EP1={final_spreads['ep1']}, "
              f"EP4={final_spreads['ep4']}, EP2={final_spreads['ep2']}, "
              f"Greek={final_spreads['greek_yes']}, Boys={final_spreads['boys']}, Girls={final_spreads['girls']}")

        return applied_swaps, final_spreads

    def _generate_all_valid_swaps(self, max_ep: int, max_gr: int, max_gen: int) -> List[Dict]:
        """
        Εξαντλητική αναζήτηση σε ΟΛΑ τα ζεύγη τμημάτων.
        Επιστρέφει μόνο swaps που περνούν τα hard rules και βελτιώνουν την ιεραρχία.
        """
        team_names = list(self.teams.keys())
        all_swaps = []

        for i, team_a in enumerate(team_names):
            for team_b in team_names[i+1:]:
                all_swaps.extend(self._generate_swaps_for_pair(team_a, team_b, max_ep, max_gr, max_gen))

        return all_swaps

    def _generate_swaps_for_pair(self, team_a: str, team_b: str,
                                  max_ep: int, max_gr: int, max_gen: int) -> List[Dict]:
        """
        Δημιουργεί νόμιμες ανταλλαγές μεταξύ δύο τμημάτων.
        Εξετάζει:
        - Solo ↔ Solo για μαθητές χωρίς πλήρως αμοιβαία δυάδα στο τμήμα τους.
        - Pair ↔ Pair για πλήρως αμοιβαίες δυάδες.
        Δεν εξετάζει Solo ↔ Pair για να μην αλλάζει αριθμούς τμημάτων.
        """
        swaps = []

        solos_a = self._get_movable_solos(team_a)
        solos_b = self._get_movable_solos(team_b)
        pairs_a = self._get_movable_pairs(team_a)
        pairs_b = self._get_movable_pairs(team_b)

        for s_a in solos_a:
            for s_b in solos_b:
                swap = self._try_swap(team_a, [s_a['name']], team_b, [s_b['name']],
                                      'Solo↔Solo', 1, max_ep, max_gr, max_gen)
                if swap:
                    swaps.append(swap)

        for p_a in pairs_a:
            for p_b in pairs_b:
                swap = self._try_swap(team_a, [p_a['name_a'], p_a['name_b']],
                                      team_b, [p_b['name_a'], p_b['name_b']],
                                      'Pair↔Pair', 2, max_ep, max_gr, max_gen)
                if swap:
                    swaps.append(swap)

        return swaps

    def _swap_creates_conflict(
        self,
        from_team: str, names_out: List[str],
        to_team: str,   names_in: List[str],
    ) -> bool:
        """
        Επιστρέφει True αν το swap θα τοποθετήσει μαθητή σε τμήμα
        που περιέχει κάποιον με δηλωμένη σύγκρουση (αμφίδρομος έλεγχος).
        """
        names_out_set = set(names_out)
        names_in_set  = set(names_in)

        def _conflicts_of(name: str) -> set:
            s = self.students.get(name)
            return {_canon_name(x) for x in s.conflict_names} if s else set()

        def _check(movers: List[str], target_team: str, leaving: set) -> bool:
            staying = [n for n in self.teams.get(target_team, []) if n not in leaving]
            for mover in movers:
                mover_c = _canon_name(mover)
                conf_mover = _conflicts_of(mover)
                for member in staying:
                    member_c = _canon_name(member)
                    if member_c in conf_mover:
                        return True
                    if mover_c in _conflicts_of(member):
                        return True
            return False

        if _check(names_out, to_team, names_in_set):
            return True
        if _check(names_in, from_team, names_out_set):
            return True
        return False

    def _try_swap(self, from_team: str, names_out: List[str],
                  to_team: str, names_in: List[str],
                  swap_type: str, priority: int,
                  max_ep: int, max_gr: int, max_gen: int) -> Optional[Dict]:
        """
        Smart-spread validation με ιεραρχία 5 → 1 → 4 → 2 → Ελληνικά → Φύλο.

        Δέχεται swap μόνο αν:
        - Δεν δημιουργεί ΣΥΓΚΡΟΥΣΗ.
        - Κρατά όλα τα βασικά spreads μέσα στα όρια ή τα μειώνει όταν είναι ήδη πάνω από όριο.
        - Δεν χειροτερεύει κανένα από τα κύρια κριτήρια.
        - Βελτιώνει τουλάχιστον ένα κριτήριο της ιεραρχίας.
        """
        spreads_before = self.calculate_spreads()
        spreads_after  = self._calc_spreads_after(from_team, names_out, to_team, names_in)

        # Hard constraint: δηλωμένη ΣΥΓΚΡΟΥΣΗ.
        if self._swap_creates_conflict(from_team, names_out, to_team, names_in):
            return None

        gender_before = max(spreads_before['boys'], spreads_before['girls'])
        gender_after  = max(spreads_after['boys'],  spreads_after['girls'])

        def _smart_ok(before: int, after: int, limit: int) -> bool:
            if before > limit:
                return after < before
            return after <= limit

        for key in self.PERFORMANCE_PRIORITY:
            if not _smart_ok(spreads_before[key], spreads_after[key], max_ep):
                return None
        if not _smart_ok(spreads_before['greek_yes'], spreads_after['greek_yes'], max_gr):
            return None
        if not _smart_ok(gender_before, gender_after, max_gen):
            return None

        deltas = {
            'delta_ep5': spreads_before['ep5'] - spreads_after['ep5'],
            'delta_ep1': spreads_before['ep1'] - spreads_after['ep1'],
            'delta_ep4': spreads_before['ep4'] - spreads_after['ep4'],
            'delta_ep2': spreads_before['ep2'] - spreads_after['ep2'],
            'delta_greek': spreads_before['greek_yes'] - spreads_after['greek_yes'],
            'delta_gender': gender_before - gender_after,
            'delta_boys': spreads_before['boys'] - spreads_after['boys'],
            'delta_girls': spreads_before['girls'] - spreads_after['girls'],
            # Συμβατότητα με παλιότερα logs/κλήσεις.
            'delta_ep3': spreads_before.get('ep3', 0) - spreads_after.get('ep3', 0),
        }

        # No harm στα κύρια κριτήρια.
        if any(deltas[k] < 0 for k in ['delta_ep5', 'delta_ep1', 'delta_ep4', 'delta_ep2', 'delta_greek', 'delta_gender']):
            return None

        # Τουλάχιστον ένα κριτήριο βελτιώνεται.
        if all(deltas[k] == 0 for k in ['delta_ep5', 'delta_ep1', 'delta_ep4', 'delta_ep2', 'delta_greek', 'delta_gender']):
            return None

        imp = {
            'improves': True,
            **deltas,
            'ep5_before': spreads_before['ep5'], 'ep5_after': spreads_after['ep5'],
            'ep1_before': spreads_before['ep1'], 'ep1_after': spreads_after['ep1'],
            'ep4_before': spreads_before['ep4'], 'ep4_after': spreads_after['ep4'],
            'ep2_before': spreads_before['ep2'], 'ep2_after': spreads_after['ep2'],
        }

        return {
            'type': swap_type,
            'from_team': from_team,
            'students_out': names_out,
            'to_team': to_team,
            'students_in': names_in,
            'improvement': imp,
            'priority': priority,
        }

    def _calc_spreads_after(self, from_team: str, names_out: List[str],
                             to_team: str, names_in: List[str]) -> Dict[str, int]:
        """Υπολογισμός spreads μετά από υποθετική ανταλλαγή."""
        stats_before = self._get_team_stats()
        stats_after = {k: v.copy() for k, v in stats_before.items()}

        def _remove(team: str, name: str) -> None:
            if name not in self.students:
                return
            s = self.students[name]
            if s.choice in {1, 2, 3, 4, 5}:
                stats_after[team][f'ep{s.choice}'] -= 1
            if s.gender == 'Α':
                stats_after[team]['boys'] -= 1
            elif s.gender == 'Κ':
                stats_after[team]['girls'] -= 1
            if s.greek_knowledge in ['Ν', 'N']:
                stats_after[team]['greek_yes'] -= 1

        def _add(team: str, name: str) -> None:
            if name not in self.students:
                return
            s = self.students[name]
            if s.choice in {1, 2, 3, 4, 5}:
                stats_after[team][f'ep{s.choice}'] += 1
            if s.gender == 'Α':
                stats_after[team]['boys'] += 1
            elif s.gender == 'Κ':
                stats_after[team]['girls'] += 1
            if s.greek_knowledge in ['Ν', 'N']:
                stats_after[team]['greek_yes'] += 1

        for name in names_out:
            _remove(from_team, name)
        for name in names_in:
            _add(from_team, name)
        for name in names_in:
            _remove(to_team, name)
        for name in names_out:
            _add(to_team, name)

        def _spread(key: str) -> int:
            vals = [s[key] for s in stats_after.values()]
            return max(vals) - min(vals) if vals else 0

        return {
            'ep5': _spread('ep5'),
            'ep1': _spread('ep1'),
            'ep4': _spread('ep4'),
            'ep2': _spread('ep2'),
            'ep3': _spread('ep3'),
            'boys': _spread('boys'),
            'girls': _spread('girls'),
            'greek_yes': _spread('greek_yes'),
        }

    def _select_best_swap_hierarchical(self, swaps: List[Dict]) -> Optional[Dict]:
        """
        Ιεραρχημένη επιλογή βέλτιστης ανταλλαγής:
        1. Μεγαλύτερη μείωση spread επίδοσης 5
        2. Μεγαλύτερη μείωση spread επίδοσης 1
        3. Μεγαλύτερη μείωση spread επίδοσης 4
        4. Μεγαλύτερη μείωση spread επίδοσης 2
        5. Μεγαλύτερη μείωση γλωσσικής επάρκειας
        6. Μεγαλύτερη μείωση φύλου
        """
        if not swaps:
            return None

        swaps.sort(key=lambda x: (
            -x['improvement']['delta_ep5'],
            -x['improvement']['delta_ep1'],
            -x['improvement']['delta_ep4'],
            -x['improvement']['delta_ep2'],
            -x['improvement']['delta_greek'],
            -x['improvement']['delta_gender'],
            x['priority']
        ))

        return swaps[0]

    def _has_mutual_friend_in_team(self, name: str, team_name: str) -> bool:
        """True αν ο μαθητής έχει πλήρως αμοιβαίο φίλο μέσα στο ίδιο τμήμα."""
        if name not in self.students:
            return False
        student = self.students[name]
        student_names = self.teams.get(team_name, [])
        team_name_by_canon = {_canon_name(n): n for n in student_names}
        for f in student.friends:
            friend_name = team_name_by_canon.get(_canon_name(f))
            if friend_name and friend_name in self.students:
                if _name_in_list(name, self.students[friend_name].friends):
                    return True
        return False

    def _get_movable_solos(self, team_name: str) -> List[Dict]:
        """Μονάδες που μπορούν να μετακινηθούν: όχι locked και χωρίς αμοιβαία δυάδα στο τμήμα."""
        solos = []
        for name in self.teams.get(team_name, []):
            if name not in self.students:
                continue
            student = self.students[name]
            if student.locked:
                continue
            if self._has_mutual_friend_in_team(name, team_name):
                continue
            solos.append({'name': name, 'student': student})
        return solos

    def _get_movable_pairs(self, team_name: str) -> List[Dict]:
        """Πλήρως αμοιβαίες δυάδες που μπορούν να μετακινηθούν μαζί: κανένας locked."""
        pairs = []
        processed = set()
        student_names = self.teams.get(team_name, [])
        for name_a in student_names:
            if name_a in processed or name_a not in self.students:
                continue
            student_a = self.students[name_a]
            if student_a.locked:
                continue
            for name_b in student_names:
                if name_b == name_a or name_b in processed or name_b not in self.students:
                    continue
                student_b = self.students[name_b]
                if student_b.locked:
                    continue
                if _are_mutual_friends(name_a, student_a.friends, name_b, student_b.friends):
                    pairs.append({
                        'name_a': name_a, 'name_b': name_b,
                        'student_a': student_a, 'student_b': student_b,
                        'ep_combo': f"{student_a.choice},{student_b.choice}"
                    })
                    processed.add(name_a)
                    processed.add(name_b)
                    break
        return pairs

    # Backward-compatible wrappers για παλιότερες κλήσεις/ονόματα.
    def _get_solos_with_ep3(self, team_name: str) -> List[Dict]:
        return [s for s in self._get_movable_solos(team_name) if s['student'].choice == 3]

    def _get_solos_without_ep3(self, team_name: str) -> List[Dict]:
        return [s for s in self._get_movable_solos(team_name) if s['student'].choice != 3]

    def _get_pairs_with_ep3(self, team_name: str) -> List[Dict]:
        return [p for p in self._get_movable_pairs(team_name) if p['student_a'].choice == 3 or p['student_b'].choice == 3]

    def _get_pairs_without_ep3(self, team_name: str) -> List[Dict]:
        return [p for p in self._get_movable_pairs(team_name) if p['student_a'].choice != 3 and p['student_b'].choice != 3]

    def _calc_asymmetric_improvement(self, team_high: str, names_out: List[str],
                                      team_low: str, names_in: List[str]) -> Dict:
        """Backward-compatible improvement calculation με τη νέα ιεραρχία."""
        before = self.calculate_spreads()
        after = self._calc_spreads_after(team_high, names_out, team_low, names_in)
        gender_before = max(before['boys'], before['girls'])
        gender_after = max(after['boys'], after['girls'])
        return {
            'improves': any(before[k] - after[k] > 0 for k in ['ep5', 'ep1', 'ep4', 'ep2', 'greek_yes']) or gender_before - gender_after > 0,
            'delta_ep5': before['ep5'] - after['ep5'],
            'delta_ep1': before['ep1'] - after['ep1'],
            'delta_ep4': before['ep4'] - after['ep4'],
            'delta_ep2': before['ep2'] - after['ep2'],
            'delta_ep3': before['ep3'] - after['ep3'],
            'delta_boys': before['boys'] - after['boys'],
            'delta_girls': before['girls'] - after['girls'],
            'delta_greek': before['greek_yes'] - after['greek_yes'],
            'delta_gender': gender_before - gender_after,
        }


    def _apply_swap(self, swap: Dict) -> None:
        from_team = swap['from_team']
        to_team = swap['to_team']
        students_out = swap['students_out']
        students_in = swap['students_in']
        
        for name in students_out:
            if name in self.teams[from_team]:
                self.teams[from_team].remove(name)
        
        for name in students_in:
            if name in self.teams[to_team]:
                self.teams[to_team].remove(name)
        
        for name in students_out:
            self.teams[to_team].append(name)
        
        for name in students_in:
            self.teams[from_team].append(name)
    
    def export_optimized_excel(self, applied_swaps: List[Dict], final_spreads: Dict, output_path: str) -> str:
        """Εξαγωγή optimized Excel."""
        wb = Workbook()
        wb.remove(wb.active)
        
        for team_name in sorted(self.teams.keys()):
            self._create_team_sheet(wb, team_name)
        
        self._create_statistics_sheet(wb, final_spreads)
        self._create_swaps_log_sheet(wb, applied_swaps)
        
        wb.save(output_path)
        wb.close()
        
        print(f"✅ Optimized Excel αποθηκεύτηκε: {output_path}")
        return output_path
    
    def _create_team_sheet(self, wb: Workbook, team_name: str) -> None:
        sheet = wb.create_sheet(team_name)
        
        headers = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDEBF7', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx = 2
        for name in sorted(self.teams[team_name]):
            if name not in self.students:
                continue
            
            student = self.students[name]
            
            # Normalize greek_knowledge
            greek_val = student.greek_knowledge
            if greek_val in ['N', 'n']:
                greek_val = 'Ν'
            elif greek_val in ['O', 'o']:
                greek_val = 'Ο'
            
            sheet.cell(row_idx, 1).value = student.name
            sheet.cell(row_idx, 2).value = student.gender
            sheet.cell(row_idx, 3).value = greek_val
            sheet.cell(row_idx, 4).value = student.choice
            sheet.cell(row_idx, 5).value = ', '.join(student.friends)
            
            for col in range(1, 6):
                sheet.cell(row_idx, col).alignment = Alignment(
                    horizontal='left' if col in [1,5] else 'center', 
                    vertical='center'
                )
            
            row_idx += 1
        
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 40
    
    def _create_statistics_sheet(self, wb: Workbook, spreads: Dict) -> None:
        sheet = wb.create_sheet('ΒΕΛΤΙΩΜΕΝΗ_ΣΤΑΤΙΣΤΙΚΗ')

        headers = ['Τμήμα', 'Σύνολο', 'Αγόρια', 'Κορίτσια',
                   'Γνώση (ΝΑΙ)', 'Γνώση (ΟΧΙ)', 'Επ1', 'Επ2', 'Επ3', 'Επ4', 'Επ5']

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='C6E0B4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        stats = self._get_team_stats()
        row_idx = 2
        for team_name in sorted(self.teams.keys()):
            if team_name not in stats:
                continue
            s = stats[team_name]

            values = [
                team_name, len(self.teams[team_name]), s['boys'], s['girls'],
                s['greek_yes'], s['greek_no'], s['ep1'], s['ep2'], s['ep3'], s['ep4'], s['ep5']
            ]
            for col_idx, value in enumerate(values, start=1):
                sheet.cell(row_idx, col_idx).value = value
                sheet.cell(row_idx, col_idx).alignment = Alignment(horizontal='center', vertical='center')

            row_idx += 1

        row_idx += 2
        sheet.cell(row_idx, 1).value = 'ΤΕΛΙΚΑ SPREADS'
        sheet.cell(row_idx, 1).font = Font(bold=True, size=12)
        row_idx += 1

        summary_headers = ['Μετρική', 'Spread', 'Στόχος', 'Status']
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = sheet.cell(row_idx, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFF2CC', fill_type='solid')
        row_idx += 1

        summary_data = [
            ('1η προτεραιότητα: Spread Επίδοσης 5', spreads['ep5'], '≤ 2', '✅' if spreads['ep5'] <= 2 else '❌'),
            ('2η προτεραιότητα: Spread Επίδοσης 1', spreads['ep1'], '≤ 2', '✅' if spreads['ep1'] <= 2 else '❌'),
            ('3η προτεραιότητα: Spread Επίδοσης 4', spreads['ep4'], '≤ 2', '✅' if spreads['ep4'] <= 2 else '❌'),
            ('4η προτεραιότητα: Spread Επίδοσης 2', spreads['ep2'], '≤ 2', '✅' if spreads['ep2'] <= 2 else '❌'),
            ('Αναφορά: Spread Επίδοσης 3', spreads.get('ep3', 0), '—', 'ℹ️'),
            ('Spread Γνώσης', spreads['greek_yes'], '≤ 2', '✅' if spreads['greek_yes'] <= 2 else '❌'),
            ('Spread Αγοριών', spreads['boys'], '≤ 2', '✅' if spreads['boys'] <= 2 else '❌'),
            ('Spread Κοριτσιών', spreads['girls'], '≤ 2', '✅' if spreads['girls'] <= 2 else '❌')
        ]

        for label, value, target, status in summary_data:
            sheet.cell(row_idx, 1).value = label
            sheet.cell(row_idx, 2).value = value
            sheet.cell(row_idx, 3).value = target
            sheet.cell(row_idx, 4).value = status

            if '✅' in status:
                sheet.cell(row_idx, 2).fill = PatternFill(start_color='C6EFCE', fill_type='solid')
            elif '❌' in status:
                sheet.cell(row_idx, 2).fill = PatternFill(start_color='FFC7CE', fill_type='solid')

            row_idx += 1

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            sheet.column_dimensions[col].width = 20
        sheet.column_dimensions['A'].width = 38

    def _create_swaps_log_sheet(self, wb: Workbook, swaps: List[Dict]) -> None:
        sheet = wb.create_sheet('ΕΦΑΡΜΟΣΜΕΝΑ_SWAPS')

        headers = ['#', 'Τύπος', 'Από Τμήμα', 'Μαθητές OUT',
                   'Προς Τμήμα', 'Μαθητές IN', 'Δ_ep5', 'Δ_ep1', 'Δ_ep4', 'Δ_ep2',
                   'Δ_γνώσης', 'Δ_φύλου', 'Priority']

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def _fmt_delta(v: int) -> str:
            return f"+{v}" if v > 0 else str(v)

        for idx, swap in enumerate(swaps, start=1):
            imp = swap['improvement']
            delta_gen = imp.get('delta_gender', imp.get('delta_boys', 0) + imp.get('delta_girls', 0))

            values = [
                idx,
                swap['type'],
                swap['from_team'],
                ', '.join(swap['students_out']),
                swap['to_team'],
                ', '.join(swap['students_in']),
                _fmt_delta(imp.get('delta_ep5', 0)),
                _fmt_delta(imp.get('delta_ep1', 0)),
                _fmt_delta(imp.get('delta_ep4', 0)),
                _fmt_delta(imp.get('delta_ep2', 0)),
                _fmt_delta(imp.get('delta_greek', 0)),
                _fmt_delta(delta_gen),
                swap['priority'],
            ]

            for col_idx, value in enumerate(values, start=1):
                sheet.cell(idx + 1, col_idx).value = value
                sheet.cell(idx + 1, col_idx).alignment = Alignment(horizontal='center', vertical='center')

        widths = [('A',8),('B',20),('C',15),('D',35),('E',15),('F',35),
                  ('G',10),('H',10),('I',10),('J',10),('K',10),('L',10),('M',10)]
        for col, width in widths:
            sheet.column_dimensions[col].width = width

    # ==================== HELPERS ====================
    
    def _parse_headers(self, sheet: Worksheet) -> Dict[str, int]:
        """Normalization headers."""
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                raw_header = str(cell.value).strip()
                headers[raw_header] = col_idx
                normalized = raw_header.upper().replace(' ', '').replace('_', '')
                headers[normalized] = col_idx
        return headers
    
    def _parse_headers_fill(self, sheet: Worksheet) -> Dict[str, int]:
        """Parse headers για fill phase."""
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                header = str(cell.value).strip()
                headers[header] = col_idx
        return headers
    
    def _get_cell_value(self, sheet: Worksheet, row: int, col: Optional[int], default: str = '') -> str:
        if col is None:
            return default
        val = sheet.cell(row, col).value
        return str(val).strip() if val is not None else default


# ========== CLI ==========

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Βήμα 8: Fill + Optimize (χωρίς Streamlit)")
    sub = p.add_subparsers(dest="mode", required=True, help="Λειτουργία")

    # Fill mode
    p_fill = sub.add_parser("fill", help="Fill template με δεδομένα")
    p_fill.add_argument("--source", required=True, help="Excel με μαθητές (Παράδειγμα1.xlsx)")
    p_fill.add_argument("--template", required=True, help="Template με τμήματα")
    p_fill.add_argument("--out", required=True, help="Output path")

    # Optimize mode
    p_opt = sub.add_parser("optimize", help="Optimize filled Excel")
    p_opt.add_argument("--input", required=True, help="Filled Excel")
    p_opt.add_argument("--out", required=True, help="Output path")
    p_opt.add_argument("--max-iter", type=int, default=100, help="Max iterations (default: 100)")
    p_opt.add_argument(
        "--allow-legacy-fallback", action="store_true",
        help=(
            "Αποδέξου input Excel από ΠΑΛΙΟ fill mode (χωρίς τις νέες raw "
            "στήλες) με fallback σε defaults, αντί να σταματήσει με σφάλμα."
        ),
    )

    # All mode
    p_all = sub.add_parser("all", help="Fill + Optimize σε μία")
    p_all.add_argument("--source", required=True)
    p_all.add_argument("--template", required=True)
    p_all.add_argument("--out", required=True)
    p_all.add_argument("--max-iter", type=int, default=100, help="Max iterations")

    return p


def main(argv: Optional[List[str]] = None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    parser = _build_parser()
    args = parser.parse_args(argv)

    try:
        processor = UnifiedProcessor()

        if args.mode == "fill":
            print(f"📄 Mode: FILL")
            processor.read_source_data(args.source)
            processor.fill_target_excel(args.template, args.out)
            
            if processor.warnings:
                print(f"\n⚠️  {len(processor.warnings)} warnings:")
                for w in processor.warnings[:10]:
                    print(f"  • {w}")
            
            return 0

        elif args.mode == "optimize":
            # FIX: πριν το mode αυτό ήταν αδρανές (μόνο print + return 1) και
            # απαιτούσε το mode 'all'. Τώρα δουλεύει standalone πάνω σε ένα
            # ήδη filled Excel (πρέπει να περιέχει τα sheets ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ
            # και/ή SINGLE, όπως τα παράγει το mode 'fill'), χωρίς να
            # χρειάζεται το αρχικό source file.
            print("📄 Mode: OPTIMIZE (standalone)")

            processor.load_filled_data(
                args.input, allow_legacy_fallback=args.allow_legacy_fallback
            )
            if not processor.students or not processor.teams:
                print(
                    "❌ Δεν βρέθηκαν δεδομένα μαθητών/τμημάτων στο input. "
                    "Βεβαιώσου ότι το αρχείο περιέχει τα sheets "
                    "ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ/SINGLE (παράγονται από το mode 'fill').",
                    file=sys.stderr,
                )
                return 1

            spreads_before = processor.calculate_spreads()
            print(f"\n📊 ΠΡΙΝ:")
            print(f"  EP3 spread: {spreads_before['ep3']}")
            print(f"  Boys spread: {spreads_before['boys']}")
            print(f"  Girls spread: {spreads_before['girls']}")
            print(f"  Greek spread: {spreads_before['greek_yes']}")

            swaps, spreads_after = processor.optimize(max_iterations=args.max_iter)

            print(f"\n📊 ΜΕΤΑ:")
            print(f"  EP3 spread: {spreads_after['ep3']} {'✅' if spreads_after['ep3'] <= 2 else '❌'}")
            print(f"  Boys spread: {spreads_after['boys']} {'✅' if spreads_after['boys'] <= 2 else '❌'}")
            print(f"  Girls spread: {spreads_after['girls']} {'✅' if spreads_after['girls'] <= 2 else '❌'}")
            print(f"  Greek spread: {spreads_after['greek_yes']} {'✅' if spreads_after['greek_yes'] <= 2 else '❌'}")

            processor.export_optimized_excel(swaps, spreads_after, args.out)
            print(f"\n🎉 Ολοκληρώθηκε! Swaps: {len(swaps)}")

            if processor.warnings:
                print(f"\n⚠️  {len(processor.warnings)} warnings:")
                for w in processor.warnings[:10]:
                    print(f"  • {w}")

            return 0

        elif args.mode == "all":
            print(f"📄 Mode: ALL (Fill + Optimize)")
            
            # Phase 1: Fill
            print("\n📋 Phase 1/2: Filling...")
            processor.read_source_data(args.source)
            temp_filled = args.out.replace('.xlsx', '_TEMP_FILLED.xlsx')
            processor.fill_target_excel(args.template, temp_filled)
            
            # Phase 2: Optimize
            print("\n🎯 Phase 2/2: Optimizing...")
            processor.load_filled_data(temp_filled)
            
            spreads_before = processor.calculate_spreads()
            print(f"\n📊 ΠΡΙΝ:")
            print(f"  EP3 spread: {spreads_before['ep3']}")
            print(f"  Boys spread: {spreads_before['boys']}")
            print(f"  Girls spread: {spreads_before['girls']}")
            print(f"  Greek spread: {spreads_before['greek_yes']}")
            
            swaps, spreads_after = processor.optimize(max_iterations=args.max_iter)
            
            print(f"\n📊 ΜΕΤΑ:")
            print(f"  EP3 spread: {spreads_after['ep3']} {'✅' if spreads_after['ep3'] <= 2 else '❌'}")
            print(f"  Boys spread: {spreads_after['boys']} {'✅' if spreads_after['boys'] <= 2 else '❌'}")
            print(f"  Girls spread: {spreads_after['girls']} {'✅' if spreads_after['girls'] <= 2 else '❌'}")
            print(f"  Greek spread: {spreads_after['greek_yes']} {'✅' if spreads_after['greek_yes'] <= 2 else '❌'}")
            
            processor.export_optimized_excel(swaps, spreads_after, args.out)
            
            # Cleanup temp file
            Path(temp_filled).unlink(missing_ok=True)
            
            print(f"\n🎉 Ολοκληρώθηκε! Swaps: {len(swaps)}")
            
            if processor.warnings:
                print(f"\n⚠️  {len(processor.warnings)} warnings:")
                for w in processor.warnings[:10]:
                    print(f"  • {w}")
            
            return 0

    except FileNotFoundError as e:
        print(f"❌ Σφάλμα: Δεν βρέθηκε αρχείο - {e}", file=sys.stderr)
        return 1
    except LegacyExcelFormatError as e:
        print(f"❌ Ασύμβατο (legacy) format Excel: {e}", file=sys.stderr)
        return 3
    except Exception as e:
        print(f"❌ Απρόσμενο σφάλμα: {e}", file=sys.stderr)
        return 2